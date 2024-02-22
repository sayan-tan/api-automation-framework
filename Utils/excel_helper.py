from pathlib import Path
import openpyxl
from Utils.json_helper import json_to_dict, obj_to_json


class ExcelHelper:

    def __init__(self, filename, sheet_name):
        path = Path(__file__).parent.parent.joinpath('Fixtures', filename)
        self.workbook = openpyxl.load_workbook(path)
        self.sheet = self.workbook[sheet_name]

    def get_row_count(self):
        total_rows = self.sheet.max_row
        return total_rows

    def get_column_count(self):
        total_columns = self.sheet.max_column
        return total_columns

    def get_column_names(self):
        col = self.get_column_count()
        li = []
        for i in range(1, col+1):
            cell = self.sheet.cell(row=1, column=i)
            li.insert(i-1, cell.value)
        return li

    def update_request_with_data(self, row_number, json_request, column_names):
        json_request = json_to_dict(json_request)
        col = self.get_column_count()
        for i in range(1, col+1):
            cell = self.sheet.cell(row=row_number, column=i)
            json_request[column_names[i-1]] = cell.value
        return obj_to_json(json_request)

    def get_cell_value_by_column(self, row_number, column_name):
        col_names = self.get_column_names()
        column_number = col_names.index(column_name)
        return self.sheet.cell(row=row_number, column=column_number+1).value
